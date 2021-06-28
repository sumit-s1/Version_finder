# -*- coding: utf-8 -*-


from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QWidget, QInputDialog, QLineEdit, QFileDialog, QPushButton
from itertools import islice
import configparser
import subprocess
import os
import fnmatch
import tkinter
import winreg
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from win32 import win32security
import string


class Ui_VersionInfo(object):

    def dropinfo(self, xcl):
        wb = openpyxl.load_workbook(xcl)
        sheet = wb.active
        i = sheet.max_row
        reg_list = []
        for cell in sheet['A']:
            reg_list.append(cell.value)
        if all(x is None for x in reg_list ):
            self.lbloutput.setText(" Input Sheet is Empty. Please try again")
            self.lbloutput.repaint()
        else:
            wb.create_sheet("Output")
            result_sheet = wb['Output']
            bold_font = Font(bold=True)
            al = Alignment(horizontal='center', vertical='center')
            color1 = 'd9ffb3'
            color2 = 'cceeff'


            result_sheet.cell(row=1, column=1).value = 'Machine'
            result_sheet.column_dimensions['A'].width = 16
            for row in result_sheet.iter_rows(min_row=1, max_row=i+1, min_col=1, max_col=1):
                for cell in row:
                    redfill = PatternFill(start_color='eeccff', end_color='eeccff', fill_type='solid')
                    cell.fill = redfill

            result_sheet.cell(row=1, column=2).value = 'R10 Drop - POS/SCO'
            result_sheet.column_dimensions['B'].width = 36
            for row in result_sheet.iter_rows(min_row=1, max_row=i+1, min_col=2, max_col=2):
                for cell in row:
                    redfill = PatternFill(start_color=color2, end_color=color2, fill_type='solid')
                    cell.fill = redfill

            result_sheet.cell(row=1, column=3).value = 'FL Drop - SCO Lane'
            result_sheet.column_dimensions['C'].width = 17
            for row in result_sheet.iter_rows(min_row=1, max_row=i+1, min_col=3, max_col=3):
                for cell in row:
                    redfill = PatternFill(start_color=color1, end_color=color1, fill_type='solid')
                    cell.fill = redfill

            result_sheet.cell(row=1, column=4).value = 'Store Server Drop'
            result_sheet.column_dimensions['D'].width = 38
            for row in result_sheet.iter_rows(min_row=1, max_row=i+1, min_col=4, max_col=4):
                for cell in row:
                    redfill = PatternFill(start_color=color2, end_color=color2, fill_type='solid')
                    cell.fill = redfill

            result_sheet.cell(row=1, column=5).value = 'POS/SCO Store Release'
            result_sheet.column_dimensions['E'].width = 33
            for row in result_sheet.iter_rows(min_row=1, max_row=i+1, min_col=5, max_col=5):
                for cell in row:
                    redfill = PatternFill(start_color=color1, end_color=color1, fill_type='solid')
                    cell.fill = redfill

            result_sheet.cell(row=1, column=6).value = 'Store Server Release'
            result_sheet.column_dimensions['F'].width = 33
            for row in result_sheet.iter_rows(min_row=1, max_row=i+1, min_col=6, max_col=6):
                for cell in row:
                    redfill = PatternFill(start_color=color2, end_color=color2, fill_type='solid')
                    cell.fill = redfill

            result_sheet.cell(row=1, column=7).value = 'FL-Store Server Patch'
            result_sheet.column_dimensions['G'].width = 37
            for row in result_sheet.iter_rows(min_row=1, max_row=i+1, min_col=7, max_col=7):
                for cell in row:
                    redfill = PatternFill(start_color=color1, end_color=color1, fill_type='solid')
                    cell.fill = redfill

            result_sheet.cell(row=1, column=8).value = 'Data Pump'
            result_sheet.column_dimensions['H'].width = 20
            for row in result_sheet.iter_rows(min_row=1, max_row=i+1, min_col=8, max_col=8):
                for cell in row:
                    redfill = PatternFill(start_color=color2, end_color=color2, fill_type='solid')
                    cell.fill = redfill

            result_sheet.cell(row=1, column=9).value = 'WKS Store Release'
            result_sheet.column_dimensions['I'].width = 20
            for row in result_sheet.iter_rows(min_row=1, max_row=i+1, min_col=9, max_col=9):
                for cell in row:
                    redfill = PatternFill(start_color=color1, end_color=color1, fill_type='solid')
                    cell.fill = redfill

            result_sheet.cell(row=1, column=10).value = 'StoreServer EFT Version'
            result_sheet.column_dimensions['J'].width = 22
            for row in result_sheet.iter_rows(min_row=1, max_row=i+1, min_col=10, max_col=10):
                for cell in row:
                    redfill = PatternFill(start_color=color2, end_color=color2, fill_type='solid')
                    cell.fill = redfill

            result_sheet.cell(row=1, column=11).value = 'POS/SCO EFT Version'
            result_sheet.column_dimensions['K'].width = 20
            for row in result_sheet.iter_rows(min_row=1, max_row=i+1, min_col=11, max_col=11):
                for cell in row:
                    redfill = PatternFill(start_color=color1, end_color=color1, fill_type='solid')
                    cell.fill = redfill

            for cell in result_sheet[1:1]:
                cell.font = bold_font
                cell.alignment = al

            for loop in range(0, i):
                register = reg_list[loop].lower()
                result_sheet.cell(row=loop + 2, column=1).value = reg_list[loop]
                self.lbloutput.setText(" Querying: " + register)
                self.lbloutput.repaint()

                #Check if system is Online
                try:
                    subprocess.check_call(['ping', '-n', '1', register], stdout=subprocess.DEVNULL,
                                          stderr=subprocess.DEVNULL)
                    is_up = True
                except subprocess.CalledProcessError:
                    is_up = False

                if is_up is True:

                    if "." in register:
                        key = register.rsplit('.', 1)[1]
                    elif "rs" in register:
                        key = "240"
                    elif "ws1" in register:
                        key = "218"
                    else:
                        key = register[-3:]
                    pos_path = "//c$//Retalix//StoreServices//POSClient//"
                    sco_path = "//c$//scot//install//InstallHistory.log"
                    rs_path = "//c$//Retalix//StoreServices//Server//PosService//Extensions//"
                    wks_path = "//c$//Windows//Build.ini"
                    storerelease_pos_path = "//c$//Windows//Version.ini"
                    storerelease_rs_path = "//d$//Upgrades//Version.ini"
                    load_path = "\c$"
                    r10 = ""
                    fl = ""
                    rs = ""
                    rg_sr = ""
                    rs_sr = ""
                    rs_fl = ""
                    dp = ""
                    wks_sr = ""
                    rs_eft = ""
                    rg_eft = ""
                    pos_full_path = register + pos_path
                    sco_full_path = register + sco_path
                    rs_full_path = register + rs_path
                    storerelease_fullpath = register + storerelease_pos_path
                    storerelease_rs_fullpath = register + storerelease_rs_path
                    wks_full_path = register + wks_path
                    full_path1 = register + load_path

                    try:

                        subprocess.call(r'net use z: /del', shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                        subprocess.call(r'net use z: \\' + full_path1 + ' /user:NSWRO\sysmoduser S3nS@tiON', shell=True,
                                        stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                    except Exception:
                        self.lbloutput.setText(register + " is inaccessible")

                    ##### R10 Version ######
                    if "rg" in register or "sc" in register or int(key) < 100:
                        try:
                            for f_name in os.listdir("//" + pos_full_path):
                                if fnmatch.fnmatch(f_name, '*WOW-POSClient-CI.txt'):
                                    version = open("//" + pos_full_path + f_name, "r")
                                    r10 = (version.readlines()[0])
                                    r10 = r10.rstrip()
                                    version.close()
                        except Exception:
                            r10 = "System Offline"
                    else:
                        r10 = "N/A"
                    result_sheet.cell(row=loop + 2, column=2).value = r10

                    ##### FL Version ######
                    if "sc" in register or 50 < int(key) < 100:
                        try:
                            sco_string = "This integration build is part of WoW FL Drop release version"
                            with open("//" + sco_full_path, "r") as file:
                                for line in file:
                                    if sco_string in line:
                                        fl = line.rsplit(':', 1)[1]
                                        fl = fl.rstrip()
                        except Exception:
                            fl = "System Offline"
                    else:
                        fl = "N/A"
                    result_sheet.cell(row=loop + 2, column=3).value = fl

                    ##### Store Sever ######
                    if "rs" in register or "240" in key:
                        try:
                            for f_name in os.listdir("//" + rs_full_path):
                                if fnmatch.fnmatch(f_name, '*WOW-StoreServer-CI.txt'):
                                    version = open("//" + rs_full_path + f_name, "r")
                                    rs = (version.readlines()[0])
                                    rs = rs.rstrip()
                                    version.close()
                        except Exception:
                            rs = "System Offline"
                    else:
                        rs = "N/A"
                    result_sheet.cell(row=loop + 2, column=4).value = rs

                    ##### POS/SCO Store Release ######
                    if "rg" in register or "sc" in register or int(key) < 100:
                        try:
                            storerelease_string = 'Full Version of this Upgrade'
                            with open("//" + storerelease_fullpath, "r") as file:
                                for line in file:
                                    if storerelease_string in line:
                                        t_string = list(islice(file, 1))[-1] + list(islice(file, 1))[-1] + \
                                                     list(islice(file, 1))[-1]
                                        t = t_string.rstrip()
                                        rg_sr = t.replace('\n', ' ')
                        except Exception:
                            rg_sr = "System Offline"
                    else:
                        rg_sr = "N/A"
                    result_sheet.cell(row=loop + 2, column=5).value = rg_sr

                    ##### RS Store Release ######
                    if "rs" in register or "240" in key:
                        try:
                            storerelease_string = 'Full Version of this Upgrade'
                            with open("//" + storerelease_rs_fullpath, "r") as file:
                                for line in file:
                                    if storerelease_string in line:
                                        tempstring = list(islice(file, 1))[-1] + list(islice(file, 1))[-1] + \
                                                     list(islice(file, 1))[
                                                         -1]
                                        t = tempstring.rstrip()
                                        rs_sr = t.replace('\n', ' ')
                        except Exception:
                            rs_sr = "System Offline"
                    else:
                        rs_sr = "N/A"
                    result_sheet.cell(row=loop + 2, column=6).value = rs_sr

                    ##### FL Store Server Patch######
                    if "rs" in register or "240" in key:
                        try:
                            patch_string = "Successfully Installed Store Server-WoW Full "
                            with open("//" + sco_full_path, "r") as file:
                                for line in file:
                                    if patch_string in line:
                                        j = (line.split(patch_string, 1)[-1])
                                        rs_fl = j.rstrip()
                        except Exception:
                            rs_fl = "System Offline"
                    else:
                        rs_fl = "N/A"
                    result_sheet.cell(row=loop + 2, column=7).value = rs_fl

                    ##### Data Pump Version ######
                    if "rs" in register or "240" in key or 'sc' in register or 50 < int(key) < 100:
                        try:
                            f = 0
                            dp_string = "Installing WOW_NCRDataPump_"
                            with open("//" + sco_full_path, "r") as file:
                                for line in file:
                                    if dp_string in line:
                                        f = 1
                                        k = line.split(dp_string, 1)[1]
                                        dp = k.rstrip()
                            if f == 0:
                                dp = "N/A"
                        except Exception:
                            dp = "System Offline"
                    else:
                        dp = "N/A"
                    result_sheet.cell(row=loop + 2, column=8).value = dp

                    ##### WKS Release Version ######
                    if "ws1" in register or "218" in key or "101" in key:
                        try:
                            w = 0
                            config = configparser.RawConfigParser()
                            config.read(r"//" + wks_full_path, encoding='utf-16')
                            a = (config.get('BUILDSTATUS', 'Release')).rstrip()
                            if a != "":
                                w = 1
                                wks_sr = a
                            if w == 0:
                                wks_sr = "System Offline"
                        except Exception:
                            wks_sr = "System Offline"
                    else:
                        wks_sr = "N/A"
                    result_sheet.cell(row=loop + 2, column=9).value = wks_sr

                    if self.eft.isChecked():
                        if "rs" in register or "240" in key:
                            try:
                                target = r"\\" + register
                                handle = win32security.LogonUser("sysmoduser", "NSWRO", "S3nS@tiON",
                                                                 win32security.LOGON32_LOGON_NEW_CREDENTIALS,
                                                                 win32security.LOGON32_PROVIDER_DEFAULT)
                                win32security.ImpersonateLoggedOnUser(handle)
                                rem_reg = winreg.ConnectRegistry(target, winreg.HKEY_LOCAL_MACHINE)
                                rem_key = winreg.OpenKey(rem_reg, r'SOFTWARE\Wow6432Node\PC-EFTPOS\LFD')
                                z = winreg.QueryValueEx(rem_key, "RelToDld")
                                rs_eft = z[0].rstrip()
                                win32security.RevertToSelf()
                                handle.Close()
                                rem_reg.Close()
                            except Exception:
                                rs_eft = "System Offline"
                        else:
                            rs_eft = "N/A"
                        result_sheet.cell(row=loop + 2, column=10).value = rs_eft

                        if "rg" in register or "sc" in register or int(key) < 100:
                            if "." in register:
                                try:
                                    z = len(key)
                                    if z == 1:
                                        key = "00" + key
                                    elif z == 2:
                                        key = "0" + key
                                    reg1 = register.rsplit('.', 1)[0] + ".240"

                                    target = r"\\" + reg1
                                    handle = win32security.LogonUser("sysmoduser", "NSWRO", "S3nS@tiON",
                                                                 win32security.LOGON32_LOGON_NEW_CREDENTIALS,
                                                                 win32security.LOGON32_PROVIDER_DEFAULT)
                                    win32security.ImpersonateLoggedOnUser(handle)
                                    rem_reg = winreg.ConnectRegistry(target, winreg.HKEY_LOCAL_MACHINE)
                                    rem_key = winreg.OpenKey(rem_reg, r'SOFTWARE\Wow6432Node\PC-EFTPOS\LFD')

                                    i = 0
                                    while True:
                                        hit = winreg.EnumKey(rem_key, i)
                                        if key in hit:
                                            new_key = r'SOFTWARE\Wow6432Node\PC-EFTPOS\LFD' + "\\" + hit
                                            rem_key = winreg.OpenKey(rem_reg, new_key)
                                            z = winreg.QueryValueEx(rem_key, "ActiveSwRelease")
                                            rg_eft = z[0].rstrip()
                                            win32security.RevertToSelf()
                                            handle.Close()
                                            rem_reg.Close()
                                            break
                                        i += 1
                                except Exception:
                                        rg_eft = "System Offline"

                            elif "rg" in register or "sc" in register:
                                try:
                                    if "rg" in register:
                                        reg1 = register.rsplit('rg', 1)[0] + "rs001"
                                    if "sc" in register:
                                        reg1 = register.rsplit('sc', 1)[0] + "rs001"
                                    target = r"\\" + reg1
                                    handle = win32security.LogonUser("sysmoduser", "NSWRO", "S3nS@tiON",
                                                                     win32security.LOGON32_LOGON_NEW_CREDENTIALS,
                                                                     win32security.LOGON32_PROVIDER_DEFAULT)
                                    win32security.ImpersonateLoggedOnUser(handle)
                                    rem_reg = winreg.ConnectRegistry(target, winreg.HKEY_LOCAL_MACHINE)
                                    rem_key = winreg.OpenKey(rem_reg, r'SOFTWARE\Wow6432Node\PC-EFTPOS\LFD')
                                    i = 0
                                    while True:
                                        hit = winreg.EnumKey(rem_key, i)
                                        if key in hit:
                                            new_key = r'SOFTWARE\Wow6432Node\PC-EFTPOS\LFD' + "\\" + hit
                                            rem_key = winreg.OpenKey(rem_reg, new_key)
                                            z = winreg.QueryValueEx(rem_key, "ActiveSwRelease")
                                            rg_eft = z[0].rstrip()
                                            win32security.RevertToSelf()
                                            handle.Close()
                                            rem_reg.Close()
                                            break
                                        i += 1
                                except Exception:
                                        rg_eft = "System Offline"
                                        result_sheet.cell(row=loop + 2, column=11).value = rg_eft
                            else:
                                rg_eft = "N/A"
                            result_sheet.cell(row=loop + 2, column=11).value = rg_eft

                    else:
                        rs_eft = "Not Required"
                        rg_eft = "Not Required"
                        result_sheet.cell(row=loop + 2, column=10).value = rs_eft
                        result_sheet.cell(row=loop + 2, column=11).value = rg_eft
                else:
                    result_sheet.cell(row=loop + 2, column=2).value = "System is Offline/Inaccessible"

                wb.save(xcl)

            self.lbloutput.setText(" Task Completed. Refer Output Sheet in the Uploaded Workbook")
            self.lbloutput.repaint()



    def setupUi(self, VersionInfo):
        VersionInfo.setObjectName("Bulk Version Finder")
        VersionInfo.resize(493, 240)
        self.centralwidget = QtWidgets.QWidget(VersionInfo)
        self.centralwidget.setObjectName("centralwidget")
        self.btnbrowse = QtWidgets.QPushButton(self.centralwidget)
        self.btnbrowse.setGeometry(QtCore.QRect(230, 110, 93, 28))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.btnbrowse.setFont(font)
        self.btnbrowse.setObjectName("btnbrowse")
        self.btnbrowse.setStyleSheet("background-color: salmon; font: bold")
        self.btnbrowse.clicked.connect(self.singlebrowse)
        self.lblinstruct = QtWidgets.QLabel(self.centralwidget)
        self.lblinstruct.setGeometry(QtCore.QRect(20, 10, 451, 61))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.lblinstruct.setFont(font)
        self.lblinstruct.setObjectName("lblinstruct")
        self.lblinstruct.setWordWrap(True)
        self.lblinstruct.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.lblinstruct.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.lblinstruct.setAlignment(QtCore.Qt.AlignCenter | QtCore.Qt.AlignVCenter)
        self.lblinstruct.setStyleSheet("background-color: mistyrose")

        self.lbloutput = QtWidgets.QLabel(self.centralwidget)
        self.lbloutput.setGeometry(QtCore.QRect(20, 160, 451, 51))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.lbloutput.setTextFormat(QtCore.Qt.RichText)
        self.lbloutput.setWordWrap(True)
        self.lbloutput.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.lbloutput.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.lbloutput.setTextInteractionFlags(QtCore.Qt.LinksAccessibleByMouse | QtCore.Qt.TextSelectableByMouse)
        self.lbloutput.setStyleSheet("background-color: yellowgreen")
        self.lbloutput.setLineWidth(13)
        self.lbloutput.setFont(font)
        self.lbloutput.setObjectName("lbloutput")

        self.eft = QtWidgets.QCheckBox(self.centralwidget)
        self.eft.setEnabled(True)
        self.eft.setGeometry(QtCore.QRect(20, 115, 171, 20))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.eft.setFont(font)
        self.eft.setChecked(False)
        self.eft.setObjectName("eft")

        VersionInfo.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(VersionInfo)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 493, 26))
        self.menubar.setObjectName("menubar")
        VersionInfo.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(VersionInfo)
        self.statusbar.setObjectName("statusbar")
        VersionInfo.setStatusBar(self.statusbar)

        self.retranslateUi(VersionInfo)
        QtCore.QMetaObject.connectSlotsByName(VersionInfo)


    def singlebrowse(self):
        try:
            filepath, _ = QtWidgets.QFileDialog.getOpenFileName(None, "Select Input File", "C:/",
                                                            "Excel (*.xlsm *.xls *.xlsx *.xlw)")
            self.dropinfo(filepath)
        except Exception:
            self.lbloutput.setText(" Invalid File. Please try again")
            self.lbloutput.repaint()


    def retranslateUi(self, VersionInfo):
        _translate = QtCore.QCoreApplication.translate
        VersionInfo.setWindowIcon(QtGui.QIcon('Version.ico'))
        VersionInfo.setWindowTitle(_translate("Batch Version Finder", "Batch Version Finder 2.0"))
        self.btnbrowse.setText(_translate("MainWindow", "Browse"))
        self.lblinstruct.setText(_translate("MainWindow", "Register Details must be saved in Column A of Active Sheet.\n"
                                                          "Check the EFT checkbox if EFT Version is also required."))
        self.lbloutput.setText(_translate("MainWindow", " Select the Input Excel Sheet"))
        self.eft.setText(_translate("MainWindow", "EFT Version Required"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    VersionInfo = QtWidgets.QMainWindow()
    ui = Ui_VersionInfo()
    ui.setupUi(VersionInfo)
    VersionInfo.show()
    sys.exit(app.exec_())
